#!/usr/bin/env python3
"""Find player-built structures in a parsed Valheim world and rank them as camera subjects.

Reads ComfyStewardView's DuckDB analytics cache directly (read-only) — no Java, no
running viewer, no re-parse of the world .db.

Method: BUILDING pieces are snapped to a fine 3-D grid; cells holding at least
--min-cell pieces are "occupied"; occupied cells touching in a 26-neighbourhood
are unioned into one cluster. That yields organic cluster shapes rather than the
fixed 500 m heatmap cells the old waypoints.json ranked, and it gives every
cluster a true 3-D bounding box because the cache stores y.

Clustering in 3-D rather than 2-D matters here: this world has builds stacked
vertically, and a flat x/z grid merges a sky platform with whatever sits on the
ground beneath it, producing clusters with nonsense 5 km bounding boxes.

Region is labelled, never filtered. Roughly 61% of this world's building pieces
sit outside the ~10.5 km world radius, in what looks like a grid of templated
build plots — heavily signed and slept in, so real, but a different subject than
the in-world settlements. Choose with --region rather than having it chosen for
you.

Outputs clusters.json (for the web app and the mod) plus clusters.csv and
clusters.xlsx (for reading, sorting, and annotating by hand).

Usage:
  python scan_clusters.py [--db PATH] [--out DIR] [--cell 16] [--min-cell 4]
                          [--min-pieces 400] [--top 0]
"""
import argparse
import csv
import json
import math
import os
import sys
from collections import defaultdict

import duckdb

DEFAULT_DB = r"C:\work\ComfyStewardView\viewer\target\ComfyEra16.duckdb"


def parse_args():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--db", default=DEFAULT_DB, help="ComfyStewardView DuckDB cache")
    p.add_argument("--out", default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "out"),
                   help="output directory")
    p.add_argument("--cell", type=float, default=16.0,
                   help="grid cell size in metres for connectivity (default 16)")
    p.add_argument("--y-cell", type=float, default=16.0,
                   help="vertical cell size in metres (default 16)")
    p.add_argument("--min-cell", type=int, default=4,
                   help="pieces required for a cell to count as occupied (default 4)")
    p.add_argument("--min-pieces", type=int, default=400,
                   help="drop clusters smaller than this (default 400)")
    p.add_argument("--top", type=int, default=0,
                   help="keep only the top N by score (0 = keep all)")
    p.add_argument("--region", default="all", choices=["all", "in-world", "outland"],
                   help="which region to keep (default all; region is always labelled)")
    p.add_argument("--world-radius", type=float, default=10500.0,
                   help="Valheim playable world radius in metres (default 10500)")
    p.add_argument("--prefab-names", default=None,
                   help="optional JSON map of prefab hash -> name (from the in-game dump)")
    p.add_argument("--world-id", default=None,
                   help="scan the newest snapshot for this world id")
    p.add_argument("--snapshot-id", type=int, default=None,
                   help="scan one exact snapshot id (takes precedence over --world-id)")
    return p.parse_args()


def table_columns(con, table):
    return {row[1] for row in con.execute(f"PRAGMA table_info('{table}')").fetchall()}


def select_snapshot(con, world_id=None, snapshot_id=None):
    columns = table_columns(con, "world_snapshot")
    if "snapshot_id" not in columns:
        if world_id or snapshot_id is not None:
            raise ValueError("legacy cache has no snapshot ids; selectors are unavailable")
        rows = con.execute(
            "SELECT NULL, source_path, parsed_at, NULL, NULL FROM world_snapshot"
        ).fetchall()
        if len(rows) != 1:
            raise ValueError(f"cache contains {len(rows)} snapshots but cannot isolate them")
        return rows[0]

    world_id_expr = "world_id" if "world_id" in columns else "NULL"
    world_name_expr = "world_name" if "world_name" in columns else "NULL"
    fields = (f"snapshot_id, source_path, parsed_at, {world_id_expr} AS world_id, "
              f"{world_name_expr} AS world_name")
    if snapshot_id is not None:
        row = con.execute(
            f"SELECT {fields} FROM world_snapshot WHERE snapshot_id = ?", [snapshot_id]
        ).fetchone()
        if not row:
            raise ValueError(f"snapshot id {snapshot_id} does not exist")
        return row
    if world_id:
        if "world_id" not in columns:
            raise ValueError("cache has no world_id column; use --snapshot-id")
        row = con.execute(
            f"SELECT {fields} FROM world_snapshot WHERE world_id = ? "
            "ORDER BY snapshot_id DESC LIMIT 1", [world_id]
        ).fetchone()
        if not row:
            raise ValueError(f"world id {world_id!r} has no snapshots")
        return row
    rows = con.execute(f"SELECT {fields} FROM world_snapshot ORDER BY snapshot_id").fetchall()
    if len(rows) != 1:
        raise ValueError(
            f"cache contains {len(rows)} snapshots; pass --world-id or --snapshot-id"
        )
    return rows[0]


class UnionFind:
    def __init__(self):
        self.parent = {}

    def add(self, x):
        self.parent.setdefault(x, x)

    def find(self, x):
        root = x
        while self.parent[root] != root:
            root = self.parent[root]
        while self.parent[x] != root:          # path compression
            self.parent[x], x = root, self.parent[x]
        return root

    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[rb] = ra


def neighbour_offsets():
    """Half of the 26-neighbourhood — the reciprocal pairs come free by symmetry."""
    out = []
    for dx in (-1, 0, 1):
        for dy in (-1, 0, 1):
            for dz in (-1, 0, 1):
                if (dx, dy, dz) > (0, 0, 0):     # lexicographic: keep one of each pair
                    out.append((dx, dy, dz))
    return out


OFFSETS = neighbour_offsets()


def build_clusters(con, cell, y_cell, min_cell):
    """Grid-snap BUILDING pieces in 3-D, then union touching occupied cells."""
    print(f"  aggregating BUILDING pieces into {cell:g}x{y_cell:g}x{cell:g} m cells ...", flush=True)
    con.execute(
        """
        CREATE OR REPLACE TEMP TABLE occupied AS
        SELECT
            CAST(floor(x / ?) AS BIGINT) AS cx,
            CAST(floor(y / ?) AS BIGINT) AS cy,
            CAST(floor(z / ?) AS BIGINT) AS cz,
            count(*)                     AS pieces
        FROM selected_zdo
        WHERE category = 'BUILDING'
        GROUP BY 1, 2, 3
        HAVING count(*) >= ?
        """,
        [cell, y_cell, cell, min_cell],
    )
    cells = con.execute("SELECT cx, cy, cz FROM occupied").fetchall()
    print(f"  {len(cells):,} occupied cells", flush=True)

    occupied = set(cells)
    uf = UnionFind()
    for key in occupied:
        uf.add(key)
    for cx, cy, cz in occupied:
        for dx, dy, dz in OFFSETS:
            nb = (cx + dx, cy + dy, cz + dz)
            if nb in occupied:
                uf.union((cx, cy, cz), nb)

    groups = defaultdict(list)
    for key in occupied:
        groups[uf.find(key)].append(key)
    print(f"  {len(groups):,} connected clusters", flush=True)
    return groups


def cluster_stats(con, cell, y_cell, groups, min_pieces):
    """Per cluster: exact 3-D box, centroid, counts, dominant prefab and creator."""
    con.execute("CREATE OR REPLACE TEMP TABLE cell_cluster "
                "(cx BIGINT, cy BIGINT, cz BIGINT, cid BIGINT)")
    rows = []
    for cid, (_root, keys) in enumerate(sorted(groups.items(), key=lambda kv: -len(kv[1])), start=1):
        for cx, cy, cz in keys:
            rows.append((cx, cy, cz, cid))
    con.executemany("INSERT INTO cell_cluster VALUES (?, ?, ?, ?)", rows)

    print("  computing 3-D extents and attribution ...", flush=True)
    con.execute(
        """
        CREATE OR REPLACE TEMP TABLE piece_cluster AS
        SELECT z.x, z.y, z.z, z.prefab_name, z.creator_id, c.cid
        FROM selected_zdo z
        JOIN cell_cluster c
          ON c.cx = CAST(floor(z.x / ?) AS BIGINT)
         AND c.cy = CAST(floor(z.y / ?) AS BIGINT)
         AND c.cz = CAST(floor(z.z / ?) AS BIGINT)
        WHERE z.category = 'BUILDING'
        """,
        [cell, y_cell, cell],
    )

    base = con.execute(
        """
        SELECT cid,
               count(*)                       AS pieces,
               min(x) AS min_x, max(x) AS max_x,
               min(y) AS min_y, max(y) AS max_y,
               min(z) AS min_z, max(z) AS max_z,
               avg(x) AS cen_x, avg(y) AS cen_y, avg(z) AS cen_z,
               median(y)                      AS med_y,
               count(DISTINCT prefab_name)    AS distinct_prefabs,
               count(DISTINCT creator_id)     AS distinct_creators
        FROM piece_cluster
        GROUP BY cid
        HAVING count(*) >= ?
        ORDER BY pieces DESC
        """,
        [min_pieces],
    ).fetchall()
    keep = {r[0] for r in base}
    print(f"  {len(keep):,} clusters at or above {min_pieces:,} pieces", flush=True)

    top_prefab = {
        cid: (pf, n)
        for cid, pf, n in con.execute(
            """
            SELECT cid, prefab_name, n FROM (
              SELECT cid, prefab_name, count(*) AS n,
                     row_number() OVER (PARTITION BY cid ORDER BY count(*) DESC) AS rk
              FROM piece_cluster GROUP BY cid, prefab_name
            ) WHERE rk = 1
            """
        ).fetchall()
    }
    top_creator = {
        cid: (cr, n)
        for cid, cr, n in con.execute(
            """
            SELECT cid, creator_id, n FROM (
              SELECT cid, creator_id, count(*) AS n,
                     row_number() OVER (PARTITION BY cid ORDER BY count(*) DESC) AS rk
              FROM piece_cluster WHERE creator_id <> 0 GROUP BY cid, creator_id
            ) WHERE rk = 1
            """
        ).fetchall()
    }

    # Landmarks inside each cluster's *footprint* — matched in x/z only, since a
    # portal or bed sits at floor level rather than in the piece cells around it.
    con.execute(
        """
        CREATE OR REPLACE TEMP TABLE footprint AS
        SELECT DISTINCT cx, cz, cid FROM cell_cluster
        """
    )
    landmarks = defaultdict(lambda: defaultdict(int))
    for cid, cat, n in con.execute(
        """
        SELECT f.cid, z.category, count(*)
        FROM selected_zdo z
        JOIN footprint f
          ON f.cx = CAST(floor(z.x / ?) AS BIGINT)
         AND f.cz = CAST(floor(z.z / ?) AS BIGINT)
        WHERE z.category IN ('PORTAL','BED','SIGN','CONTAINER','ITEM_STAND','INTERIOR')
        GROUP BY 1, 2
        """,
        [cell, cell],
    ).fetchall():
        landmarks[cid][cat] = n

    return base, top_prefab, top_creator, landmarks


def score(pieces, height, footprint_m2, distinct_prefabs, cells):
    """Heuristic 'worth photographing' score. Deliberately transparent, not tuned.

    Rewards mass, vertical relief, and material variety; penalises sprawl, because a
    field of scattered fence posts covers ground without ever filling a frame.
    """
    mass = math.log10(max(pieces, 1))
    relief = min(height, 60.0) / 20.0
    variety = min(distinct_prefabs, 60) / 30.0
    density = pieces / max(footprint_m2, 1.0)          # pieces per m^2 of footprint
    compactness = min(density * 4.0, 2.0)
    return round(mass * 2.0 + relief + variety + compactness, 3)


def main():
    args = parse_args()
    if not os.path.exists(args.db):
        sys.exit(f"DuckDB cache not found: {args.db}")
    os.makedirs(args.out, exist_ok=True)

    names = {}
    if args.prefab_names and os.path.exists(args.prefab_names):
        with open(args.prefab_names, encoding="utf-8") as fh:
            names = {str(k): v for k, v in json.load(fh).items()}

    print(f"opening {args.db} (read-only)", flush=True)
    con = duckdb.connect(args.db, read_only=True)
    try:
        snap = select_snapshot(con, args.world_id, args.snapshot_id)
    except ValueError as exc:
        sys.exit(str(exc))
    snapshot_id, source_path, parsed_at, world_id, world_name = snap
    zdo_columns = table_columns(con, "zdo")
    if snapshot_id is not None and "snapshot_id" in zdo_columns:
        con.execute(f"CREATE TEMP VIEW selected_zdo AS "
                    f"SELECT * FROM zdo WHERE snapshot_id = {int(snapshot_id)}")
    elif snapshot_id is not None:
        snapshot_count = con.execute("SELECT count(*) FROM world_snapshot").fetchone()[0]
        if snapshot_count != 1:
            sys.exit("zdo table has no snapshot_id; cannot isolate a multi-snapshot cache")
        con.execute("CREATE TEMP VIEW selected_zdo AS SELECT * FROM zdo")
    else:
        con.execute("CREATE TEMP VIEW selected_zdo AS SELECT * FROM zdo")
    snapshot_label = world_name or world_id or os.path.basename(source_path)
    print(f"  snapshot {snapshot_id if snapshot_id is not None else 'legacy'}: "
          f"{snapshot_label}", flush=True)

    groups = build_clusters(con, args.cell, args.y_cell, args.min_cell)
    base, top_prefab, top_creator, landmarks = cluster_stats(
        con, args.cell, args.y_cell, groups, args.min_pieces)

    clusters = []
    for (cid, pieces, min_x, max_x, min_y, max_y, min_z, max_z,
         cen_x, cen_y, cen_z, med_y, distinct_prefabs, distinct_creators) in base:
        size_x, size_y, size_z = max_x - min_x, max_y - min_y, max_z - min_z
        footprint = max(size_x, 1.0) * max(size_z, 1.0)
        diag = math.sqrt(size_x ** 2 + size_y ** 2 + size_z ** 2)
        pf, pf_n = top_prefab.get(cid, (None, 0))
        cr, cr_n = top_creator.get(cid, (None, 0))
        lm = landmarks.get(cid, {})
        radius = math.sqrt(cen_x ** 2 + cen_z ** 2)
        region = "in-world" if radius <= args.world_radius else "outland"
        clusters.append({
            "cluster_id": cid,
            "region": region,
            "radius_m": round(radius, 1),
            "sky": med_y > 500,
            "pieces": pieces,
            "score": score(pieces, size_y, footprint, distinct_prefabs, len(groups)),
            # where to stand: centre of the footprint, at the height most pieces sit at
            "center_x": round(cen_x, 1),
            "center_y": round(med_y, 1),
            "center_z": round(cen_z, 1),
            "min_x": round(min_x, 1), "max_x": round(max_x, 1),
            "min_y": round(min_y, 1), "max_y": round(max_y, 1),
            "min_z": round(min_z, 1), "max_z": round(max_z, 1),
            "size_x": round(size_x, 1),
            "size_y": round(size_y, 1),
            "size_z": round(size_z, 1),
            "footprint_m2": round(footprint, 1),
            "diagonal_m": round(diag, 1),
            # a camera this far back, this high, frames the whole box at ~60 deg FOV
            "suggested_standoff_m": round(max(diag * 0.9, 20.0), 1),
            "suggested_camera_y": round(med_y + max(size_y * 0.6, 8.0), 1),
            "distinct_prefabs": distinct_prefabs,
            "distinct_creators": distinct_creators,
            "top_prefab": names.get(str(pf), pf),
            "top_prefab_share": round(pf_n / pieces, 3) if pieces else 0.0,
            "top_creator_id": cr,
            "top_creator_share": round(cr_n / pieces, 3) if pieces and cr_n else 0.0,
            "portals": lm.get("PORTAL", 0),
            "beds": lm.get("BED", 0),
            "signs": lm.get("SIGN", 0),
            "containers": lm.get("CONTAINER", 0),
            "item_stands": lm.get("ITEM_STAND", 0),
            "interiors": lm.get("INTERIOR", 0),
            # filled in by hand while flying
            "visited": "",
            "verdict": "",
            "notes": "",
        })

    if args.region != "all":
        clusters = [c for c in clusters if c["region"] == args.region]

    clusters.sort(key=lambda c: -c["score"])
    for rank, c in enumerate(clusters, start=1):
        c["rank"] = rank
    if args.top:
        clusters = clusters[: args.top]

    fields = ["rank", "cluster_id", "region", "sky", "radius_m", "score", "pieces",
              "center_x", "center_y", "center_z",
              "size_x", "size_y", "size_z", "footprint_m2", "diagonal_m",
              "suggested_standoff_m", "suggested_camera_y",
              "min_x", "max_x", "min_y", "max_y", "min_z", "max_z",
              "distinct_prefabs", "distinct_creators",
              "top_prefab", "top_prefab_share", "top_creator_id", "top_creator_share",
              "portals", "beds", "signs", "containers", "item_stands", "interiors",
              "visited", "verdict", "notes"]

    doc = {
        "world": world_name or world_id or os.path.basename(source_path),
        "world_id": world_id,
        "snapshot_id": snapshot_id,
        "parsed_at": str(parsed_at),
        "method": (f"3-D connected-component clustering of BUILDING ZDOs on a "
                   f"{args.cell:g}x{args.y_cell:g}x{args.cell:g} m grid "
                   f"(min {args.min_cell} pieces/cell, min {args.min_pieces} pieces/cluster, "
                   f"world radius {args.world_radius:g} m, region={args.region})"),
        "prefab_names_resolved": bool(names),
        "count": len(clusters),
        "clusters": clusters,
    }

    json_path = os.path.join(args.out, "clusters.json")
    with open(json_path, "w", encoding="utf-8") as fh:
        json.dump(doc, fh, indent=2, ensure_ascii=False)

    csv_path = os.path.join(args.out, "clusters.csv")
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for c in clusters:
            w.writerow({k: c.get(k, "") for k in fields})

    xlsx_path = write_xlsx(os.path.join(args.out, "clusters.xlsx"), clusters, fields, doc)

    print()
    print(f"wrote {len(clusters):,} clusters")
    for path in (json_path, csv_path, xlsx_path):
        if path:
            print(f"  {path}")
    print()
    n_in = sum(1 for c in clusters if c["region"] == "in-world")
    print(f"  region: {n_in:,} in-world, {len(clusters) - n_in:,} outland")
    print()
    print(f"  {'#':>4} {'region':>9} {'score':>6} {'pieces':>8} {'h':>5} {'foot m2':>9}  centre")
    for c in clusters[:15]:
        print(f"  {c['rank']:>4} {c['region']:>9} {c['score']:>6} {c['pieces']:>8,} "
              f"{c['size_y']:>5.0f} {c['footprint_m2']:>9,.0f}  "
              f"({c['center_x']:.0f}, {c['center_y']:.0f}, {c['center_z']:.0f})")


def write_xlsx(path, clusters, fields, doc):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  (openpyxl not installed — skipping .xlsx)")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "clusters"
    ws.append(fields)
    head = ws[1]
    fill = PatternFill("solid", fgColor="1F3A5F")
    for cell in head:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in clusters:
        ws.append([c.get(k, "") for k in fields])

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(clusters) + 1}"
    for i, name in enumerate(fields, start=1):
        width = 11
        if name in ("top_prefab", "notes", "verdict"):
            width = 22
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30

    meta = wb.create_sheet("about")
    for row in (("world", doc["world"]), ("parsed_at", doc["parsed_at"]),
                ("method", doc["method"]), ("clusters", doc["count"]),
                ("prefab_names_resolved", doc["prefab_names_resolved"]),
                ("", ""),
                ("score", "heuristic: 2*log10(pieces) + relief + variety + compactness"),
                ("center_x/y/z", "centroid x/z at median piece height — the teleport target"),
                ("suggested_standoff_m", "camera distance that frames the whole box at ~60 deg FOV"),
                ("visited/verdict/notes", "yours to fill in while flying")):
        meta.append(list(row))
    meta.column_dimensions["A"].width = 24
    meta.column_dimensions["B"].width = 78
    for cell in meta["A"]:
        cell.font = Font(bold=True)

    wb.save(path)
    return path


if __name__ == "__main__":
    main()
